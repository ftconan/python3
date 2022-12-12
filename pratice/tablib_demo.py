"""
# @author: magician
# @file:   tablib_demo.py
# @date:   2022/1/24
"""
import tablib
import xlrd
import copy
import openpyxl

from xlutils.copy import copy


def tablib_export_excel():
    """
    tablib_export_excel
    """
    excel_path = '/Users/magician/Projects/github/python3/data/上海备案模版.xlsx'
    with open(excel_path, 'rb') as fh:
        excel_data = tablib.Dataset().load(fh, 'xlsx', headers=False)

    data_set = tablib.Dataset(excel_data.dict[1])
    data_set.append('*'*len(excel_data.dict[1]))
    export_excel_path = '/Users/magician/Projects/github/python3/data/shanghai_meton.xlsx'
    with open(export_excel_path, 'wb+') as f:
        f.write(data_set.export('xlsx'))
    print(f'export_excel_path: {export_excel_path}')


def xlrd_export_excel():
    """
    xlrd_export_excel
    """
    # xlrd.biffh.XLRDError: Excel xlsx file; not supported
    excel_path = '/Users/magician/Projects/github/python3/data/上海备案模版.xls'
    workbook = xlrd.open_workbook(excel_path, formatting_info=True)
    wb = copy(workbook)
    sheet = wb.get_sheet('场所信息导入')
    sheet.write(2, 3, 'A3')
    export_excel_path = '/Users/magician/Projects/github/python3/data/shanghai_meton1.xls'
    wb.save(export_excel_path)
    print(f'export_excel_path: {export_excel_path}')


def openpyxl_export_excel():
    """
    openpyxl_export_excel
    """
    excel_path = '/Users/magician/Projects/github/python3/data/上海备案模版.xlsx'
    wb = openpyxl.load_workbook(excel_path)

    sheet = wb[wb.sheetnames[0]]
    for row in sheet.iter_rows(min_row=3, max_col=3, max_row=4):
        for cell in row:
            cell.value = '1'

    # sheet['A3'] = 'A3'
    export_excel_path = '/Users/magician/Projects/github/python3/data/shanghai_meton.xlsx'
    wb.save(export_excel_path)
    print(f'export_excel_path: {export_excel_path}')


if __name__ == '__main__':
    # tablib_export_excel()
    # xlrd_export_excel()
    openpyxl_export_excel()
